"""Tests for ReportExporter - Excel and PDF report generation."""

import pytest
import pandas as pd
from pathlib import Path
from tempfile import TemporaryDirectory
from openpyxl import load_workbook

from src.analyzer import DataAnalyzer
from src.yearly_analyzer import YearlyAnalyzer
from src.forecaster import Forecaster
from src.anomaly_detector import AnomalyDetector
from src.report_exporter import ReportExporter


@pytest.fixture
def sample_yearly_dataframe():
    """Create 12 months of sample invoice data."""
    dfs = []
    for month in range(1, 13):
        df = pd.DataFrame({
            'MANDT': [1] * 100,
            'FACTURA': [f'001-{month:02d}-' + str(i).zfill(6) for i in range(100)],
            'RUBRO': ['AG01'] * 60 + ['AL01'] * 40,
            'SECUENCIA': range(1, 101),
            'BLOQUE_FACTURA': ['A'] * 100,
            'ID_SUBTOTAL': range(1, 101),
            'DESC_RUBRO': ['Agua'] * 60 + ['Alcantarillado'] * 40,
            'PRECIO_UNI': [2.0] * 100,
            'PRECIO_TOTAL': [10.0 + (month * 0.5)] * 100,
            'PRECIO_DESC': [0.0] * 100,
            'CANTIDAD': [5.0] * 100,
            'MONTO_IVA': [2.0] * 100,
            'TARIFA': [1] * 100,
            'CONSU_HID': [''] * 100,
            'MONTO_NEG': [''] * 100,
        })
        dfs.append(df)
    return dfs


@pytest.fixture
def yearly_analyzer(sample_yearly_dataframe):
    """Create YearlyAnalyzer with 12 months of data."""
    month_data = {
        str(i).zfill(2): DataAnalyzer(df)
        for i, df in enumerate(sample_yearly_dataframe, 1)
    }
    return YearlyAnalyzer(month_data)


@pytest.fixture
def exporter_full(yearly_analyzer):
    """Create ReportExporter with forecast and anomaly detection."""
    forecaster = Forecaster(yearly_analyzer)
    detector = AnomalyDetector(yearly_analyzer)
    return ReportExporter(yearly_analyzer, forecaster, detector)


@pytest.fixture
def exporter_minimal(yearly_analyzer):
    """Create ReportExporter without forecast/anomaly sections."""
    return ReportExporter(yearly_analyzer)


class TestReportExporterInit:
    """Test ReportExporter initialization."""

    def test_init_with_yearly_analyzer(self, yearly_analyzer):
        """Test initialization with just YearlyAnalyzer."""
        exporter = ReportExporter(yearly_analyzer)
        assert exporter.analyzer is not None
        assert exporter.forecaster is None
        assert exporter.detector is None

    def test_init_with_all_components(self, exporter_full):
        """Test initialization with forecaster and detector."""
        assert exporter_full.forecaster is not None
        assert exporter_full.detector is not None

    def test_init_invalid_type(self):
        """Test initialization with invalid type raises error."""
        with pytest.raises(TypeError):
            ReportExporter("invalid")


class TestExcelExport:
    """Test Excel report export."""

    def test_export_excel_creates_file(self, exporter_full):
        """Test that Excel file is created."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            result = exporter_full.export_excel(filepath)

            assert Path(result).exists()
            assert result == filepath

    def test_export_excel_creates_parent_dirs(self, exporter_full):
        """Test that parent directories are created if missing."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/nested/dir/report.xlsx"
            exporter_full.export_excel(filepath)

            assert Path(filepath).exists()

    def test_export_excel_has_all_sheets(self, exporter_full):
        """Test that Excel workbook has all expected sheets."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            exporter_full.export_excel(filepath)

            wb = load_workbook(filepath)
            expected_sheets = ["Summary", "Monthly Detail", "Categories", "Forecast", "Anomalies"]
            for sheet in expected_sheets:
                assert sheet in wb.sheetnames

    def test_export_excel_minimal_skips_optional_sheets(self, exporter_minimal):
        """Test that Excel without forecaster/detector skips those sheets."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            exporter_minimal.export_excel(filepath)

            wb = load_workbook(filepath)
            assert "Summary" in wb.sheetnames
            assert "Forecast" not in wb.sheetnames
            assert "Anomalies" not in wb.sheetnames

    def test_excel_summary_sheet_has_data(self, exporter_full):
        """Test Summary sheet contains expected metric labels."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            exporter_full.export_excel(filepath)

            wb = load_workbook(filepath)
            ws = wb["Summary"]

            labels = [ws.cell(row=r, column=1).value for r in range(1, 20)]
            assert "Total Revenue" in labels
            assert "Peak Month" in labels

    def test_excel_monthly_sheet_has_all_months(self, exporter_full):
        """Test Monthly Detail sheet has a row per month."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            exporter_full.export_excel(filepath)

            wb = load_workbook(filepath)
            ws = wb["Monthly Detail"]

            months_in_sheet = [ws.cell(row=r, column=1).value for r in range(2, 14)]
            assert months_in_sheet.count(None) == 0

    def test_excel_categories_sheet_has_data(self, exporter_full):
        """Test Categories sheet has category rows."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.xlsx"
            exporter_full.export_excel(filepath)

            wb = load_workbook(filepath)
            ws = wb["Categories"]

            rubros = [ws.cell(row=r, column=1).value for r in range(2, 12)]
            assert "AG01" in rubros or "AL01" in rubros


class TestPdfExport:
    """Test PDF report export."""

    def test_export_pdf_creates_file(self, exporter_full):
        """Test that PDF file is created."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.pdf"
            result = exporter_full.export_pdf(filepath)

            assert Path(result).exists()
            assert result == filepath

    def test_export_pdf_creates_parent_dirs(self, exporter_full):
        """Test that parent directories are created if missing."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/nested/dir/report.pdf"
            exporter_full.export_pdf(filepath)

            assert Path(filepath).exists()

    def test_export_pdf_nonempty(self, exporter_full):
        """Test that PDF file has non-trivial size."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.pdf"
            exporter_full.export_pdf(filepath)

            assert Path(filepath).stat().st_size > 1000

    def test_export_pdf_minimal_still_works(self, exporter_minimal):
        """Test PDF export works without forecaster/detector."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.pdf"
            result = exporter_minimal.export_pdf(filepath)

            assert Path(result).exists()

    def test_export_pdf_custom_title(self, exporter_full):
        """Test PDF export accepts custom title."""
        with TemporaryDirectory() as tmpdir:
            filepath = f"{tmpdir}/report.pdf"
            result = exporter_full.export_pdf(filepath, title="Custom Report Title")

            assert Path(result).exists()


class TestFromDirectory:
    """Test ReportExporter.from_directory() factory."""

    def test_from_directory_loads_data(self, sample_yearly_dataframe):
        """Test loading from directory with real files."""
        with TemporaryDirectory() as tmpdir:
            for i, df in enumerate(sample_yearly_dataframe, 1):
                df.to_csv(f"{tmpdir}/Datalle {i:02d}25.txt", sep='|', index=False)

            exporter = ReportExporter.from_directory(tmpdir)
            assert exporter.analyzer is not None
            assert exporter.forecaster is not None
            assert exporter.detector is not None

    def test_from_directory_skip_forecast(self, sample_yearly_dataframe):
        """Test loading without forecaster."""
        with TemporaryDirectory() as tmpdir:
            for i, df in enumerate(sample_yearly_dataframe, 1):
                df.to_csv(f"{tmpdir}/Datalle {i:02d}25.txt", sep='|', index=False)

            exporter = ReportExporter.from_directory(tmpdir, include_forecast=False)
            assert exporter.forecaster is None

    def test_from_directory_skip_anomalies(self, sample_yearly_dataframe):
        """Test loading without anomaly detector."""
        with TemporaryDirectory() as tmpdir:
            for i, df in enumerate(sample_yearly_dataframe, 1):
                df.to_csv(f"{tmpdir}/Datalle {i:02d}25.txt", sep='|', index=False)

            exporter = ReportExporter.from_directory(tmpdir, include_anomalies=False)
            assert exporter.detector is None


class TestIntegration:
    """Integration tests combining full workflow."""

    def test_full_export_workflow(self, sample_yearly_dataframe):
        """Test complete workflow from directory to both exports."""
        with TemporaryDirectory() as tmpdir:
            for i, df in enumerate(sample_yearly_dataframe, 1):
                df.to_csv(f"{tmpdir}/Datalle {i:02d}25.txt", sep='|', index=False)

            exporter = ReportExporter.from_directory(tmpdir)

            excel_path = exporter.export_excel(f"{tmpdir}/out/report.xlsx")
            pdf_path = exporter.export_pdf(f"{tmpdir}/out/report.pdf")

            assert Path(excel_path).exists()
            assert Path(pdf_path).exists()
