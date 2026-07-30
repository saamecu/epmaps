"""Report export module for generating Excel and PDF reports from yearly analysis."""

from typing import Dict, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

from src.yearly_analyzer import YearlyAnalyzer
from src.forecaster import Forecaster
from src.anomaly_detector import AnomalyDetector


HEADER_FILL = "3B82F6"
ACCENT_FILL = "06B6D4"
LIGHT_FILL = "F8FAFC"

TREND_LABELS = {
    "increasing": "CRECIENTE",
    "decreasing": "DECRECIENTE",
    "stable": "ESTABLE",
}

SEVERITY_LABELS = {
    "high": "Alto",
    "medium": "Medio",
    "low": "Bajo",
}

ALERT_TYPE_LABELS = {
    "revenue": "Ingresos",
    "price": "Precio",
    "category": "Categoría",
}


class ReportExporter:
    """Exports yearly analysis, forecasts, and anomalies to Excel and PDF.

    Combines data from YearlyAnalyzer, Forecaster, and AnomalyDetector
    into professional, shareable report documents (in Spanish).
    """

    def __init__(
        self,
        yearly_analyzer: YearlyAnalyzer,
        forecaster: Optional[Forecaster] = None,
        anomaly_detector: Optional[AnomalyDetector] = None,
    ) -> None:
        """Initialize ReportExporter.

        Args:
            yearly_analyzer: YearlyAnalyzer instance with loaded data.
            forecaster: Optional Forecaster for forecast sections.
            anomaly_detector: Optional AnomalyDetector for anomaly sections.
        """
        if not isinstance(yearly_analyzer, YearlyAnalyzer):
            raise TypeError("yearly_analyzer must be a YearlyAnalyzer instance")

        self.analyzer = yearly_analyzer
        self.forecaster = forecaster
        self.detector = anomaly_detector

    @classmethod
    def from_directory(cls, directory: str, include_forecast: bool = True, include_anomalies: bool = True):
        """Create ReportExporter from directory of data files.

        Args:
            directory: Path to directory with monthly data files.
            include_forecast: Load Forecaster for forecast sections (default: True).
            include_anomalies: Load AnomalyDetector for anomaly sections (default: True).

        Returns:
            ReportExporter instance.
        """
        analyzer = YearlyAnalyzer.from_directory(directory)
        forecaster = Forecaster(analyzer) if include_forecast else None
        detector = AnomalyDetector(analyzer) if include_anomalies else None
        return cls(analyzer, forecaster, detector)

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------

    def export_excel(self, filepath: str) -> str:
        """Export comprehensive report to Excel workbook.

        Creates multiple sheets: Resumen, Detalle Mensual, Categorías,
        Pronóstico (if available), Anomalías (if available).

        Args:
            filepath: Path to save the .xlsx file.

        Returns:
            Path to the saved file.
        """
        wb = Workbook()
        wb.remove(wb.active)

        self._build_summary_sheet(wb)
        self._build_monthly_sheet(wb)
        self._build_categories_sheet(wb)

        if self.forecaster is not None:
            self._build_forecast_sheet(wb)

        if self.detector is not None:
            self._build_anomalies_sheet(wb)

        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        return filepath

    def _style_header_row(self, ws, row: int, n_cols: int) -> None:
        """Apply header styling to a row."""
        fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit_columns(self, ws, n_cols: int, width: int = 18) -> None:
        """Set uniform column widths."""
        for col in range(1, n_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _build_summary_sheet(self, wb: Workbook) -> None:
        """Build the Resumen sheet with yearly metrics."""
        ws = wb.create_sheet("Resumen")
        metrics = self.analyzer.yearly_metrics()

        ws["A1"] = "EPMaps - Análisis Anual - Resumen"
        ws["A1"].font = Font(bold=True, size=16, color=HEADER_FILL)
        ws.merge_cells("A1:B1")

        rows = [
            ("Ingresos Totales", f"${metrics['total_revenue']:,.0f}"),
            ("Ingreso Promedio Mensual", f"${metrics['avg_monthly_revenue']:,.0f}"),
            ("Desviación Estándar Mensual", f"${metrics['std_monthly_revenue']:,.0f}"),
            ("Coeficiente de Variación", f"{metrics['cv_monthly_revenue']:.1f}%"),
            ("Ingreso Mensual Mínimo", f"${metrics['min_monthly_revenue']:,.0f}"),
            ("Ingreso Mensual Máximo", f"${metrics['max_monthly_revenue']:,.0f}"),
            ("Mes Pico", metrics['peak_month']),
            ("Mes Bajo", metrics['low_month']),
            ("Registros Totales", f"{metrics['total_records']:,}"),
            ("Facturas Totales", f"{metrics['total_invoices']:,}"),
            ("Líneas Promedio por Factura", f"{metrics['avg_lines_per_invoice']:.2f}"),
        ]

        start_row = 3
        ws.cell(row=start_row, column=1, value="Métrica")
        ws.cell(row=start_row, column=2, value="Valor")
        self._style_header_row(ws, start_row, 2)

        for i, (label, value) in enumerate(rows, start=start_row + 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
            if i % 2 == 0:
                for col in (1, 2):
                    ws.cell(row=i, column=col).fill = PatternFill(
                        start_color=LIGHT_FILL, end_color=LIGHT_FILL, fill_type="solid"
                    )

        self._autofit_columns(ws, 2, width=30)

    def _build_monthly_sheet(self, wb: Workbook) -> None:
        """Build the Detalle Mensual sheet with a revenue chart."""
        ws = wb.create_sheet("Detalle Mensual")
        metrics = self.analyzer.yearly_metrics()['monthly_metrics']
        changes = self.analyzer.monthly_changes()

        headers = ["Mes", "Ingresos", "Registros", "Facturas", "Precio Promedio", "Cambio Mensual %"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))

        for i, month in enumerate(self.analyzer.months, start=2):
            m = metrics[month]
            change = changes[month].get('revenue_change')
            ws.cell(row=i, column=1, value=month)
            ws.cell(row=i, column=2, value=round(m['revenue'], 2))
            ws.cell(row=i, column=3, value=m['records'])
            ws.cell(row=i, column=4, value=m['invoices'])
            ws.cell(row=i, column=5, value=round(m['avg_price'], 2))
            ws.cell(row=i, column=6, value=change if change is not None else "—")

        self._autofit_columns(ws, len(headers))

        # Revenue trend chart
        n_months = len(self.analyzer.months)
        chart = LineChart()
        chart.title = "Tendencia de Ingresos Mensuales"
        chart.y_axis.title = "Ingresos ($)"
        chart.x_axis.title = "Mes"

        data = Reference(ws, min_col=2, min_row=1, max_row=n_months + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_months + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18

        ws.add_chart(chart, f"H2")

    def _build_categories_sheet(self, wb: Workbook) -> None:
        """Build the Categorías sheet with top category breakdown."""
        ws = wb.create_sheet("Categorías")
        top_cats = self.analyzer.top_categories_yearly(top_n=10)

        headers = ["Categoría", "Ingresos Totales", "% del Total", "CV %", "Mes Mínimo", "Mes Máximo", "Registros"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))

        for i, (rubro, data) in enumerate(top_cats.items(), start=2):
            ws.cell(row=i, column=1, value=rubro)
            ws.cell(row=i, column=2, value=round(data['total_revenue'], 2))
            ws.cell(row=i, column=3, value=data['pct_of_total'])
            ws.cell(row=i, column=4, value=data['cv_monthly'])
            ws.cell(row=i, column=5, value=data['min_month'])
            ws.cell(row=i, column=6, value=data['max_month'])
            ws.cell(row=i, column=7, value=data['records_total'])

        self._autofit_columns(ws, len(headers))

        # Category revenue bar chart
        n_cats = len(top_cats)
        chart = BarChart()
        chart.title = "Ingresos por Categoría"
        chart.y_axis.title = "Ingresos ($)"
        chart.x_axis.title = "Categoría"

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=n_cats + 1)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=n_cats + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 8
        chart.width = 18

        ws.add_chart(chart, "I2")

    def _build_forecast_sheet(self, wb: Workbook) -> None:
        """Build the Pronóstico sheet with predictions."""
        ws = wb.create_sheet("Pronóstico")
        forecast = self.forecaster.get_forecast_summary(periods=3)
        trends = self.forecaster.get_trends()

        ws["A1"] = "Pronóstico Q1 y Análisis de Tendencia"
        ws["A1"].font = Font(bold=True, size=16, color=HEADER_FILL)
        ws.merge_cells("A1:D1")

        ws["A3"] = "Tipo de Tendencia"
        ws["B3"] = TREND_LABELS.get(trends['trend_type'], trends['trend_type'].upper())
        ws["A4"] = "Cambio Anual %"
        ws["B4"] = f"{trends['annual_pct_change']:.1f}%"
        ws["A5"] = "Volatilidad (CV)"
        ws["B5"] = f"{trends['volatility_cv']:.1f}%"
        ws["A6"] = "Fuerza de Estacionalidad"
        ws["B6"] = f"{trends['seasonality_strength']:.1f}%"

        for row in range(3, 7):
            ws.cell(row=row, column=1).font = Font(bold=True)

        headers = ["Mes", "Pronóstico", "Límite Inferior", "Límite Superior"]
        start_row = 8
        for col, header in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col, value=header)
        self._style_header_row(ws, start_row, len(headers))

        for i, (month, data) in enumerate(forecast['revenue_forecast'].items(), start=start_row + 1):
            ws.cell(row=i, column=1, value=month)
            ws.cell(row=i, column=2, value=round(data['forecast'], 2))
            ws.cell(row=i, column=3, value=round(data['lower_bound'], 2))
            ws.cell(row=i, column=4, value=round(data['upper_bound'], 2))

        self._autofit_columns(ws, len(headers), width=20)

    def _build_anomalies_sheet(self, wb: Workbook) -> None:
        """Build the Anomalías sheet with detection results."""
        ws = wb.create_sheet("Anomalías")
        report = self.detector.get_overall_anomaly_report()

        ws["A1"] = "Reporte de Detección de Anomalías"
        ws["A1"].font = Font(bold=True, size=16, color=HEADER_FILL)
        ws.merge_cells("A1:D1")

        ws["A3"] = "Nivel de Riesgo"
        ws["B3"] = SEVERITY_LABELS.get(report['risk_level'], report['risk_level'].upper())
        ws["A4"] = "Puntuación de Riesgo"
        ws["B4"] = f"{report['risk_score']:.1f}/100"
        ws["A5"] = "Anomalías Totales"
        ws["B5"] = report['total_anomalies_detected']

        for row in range(3, 6):
            ws.cell(row=row, column=1).font = Font(bold=True)

        headers = ["Tipo", "Mes", "Valor", "Z-Score", "Desviación %", "Severidad"]
        start_row = 7
        for col, header in enumerate(headers, start=1):
            ws.cell(row=start_row, column=col, value=header)
        self._style_header_row(ws, start_row, len(headers))

        row_idx = start_row + 1
        for month, anom in report['revenue_anomalies'].get('anomalies', {}).items():
            ws.cell(row=row_idx, column=1, value="Ingresos")
            ws.cell(row=row_idx, column=2, value=month)
            ws.cell(row=row_idx, column=3, value=round(anom['revenue'], 2))
            ws.cell(row=row_idx, column=4, value=round(anom['z_score'], 2))
            ws.cell(row=row_idx, column=5, value=round(anom['deviation_from_mean_pct'], 1))
            ws.cell(row=row_idx, column=6, value=SEVERITY_LABELS.get(anom['severity'], anom['severity']))
            row_idx += 1

        for month, anom in report['price_anomalies'].get('anomalies', {}).items():
            ws.cell(row=row_idx, column=1, value="Precio")
            ws.cell(row=row_idx, column=2, value=month)
            ws.cell(row=row_idx, column=3, value=round(anom['avg_price'], 2))
            ws.cell(row=row_idx, column=4, value=round(anom['z_score'], 2))
            ws.cell(row=row_idx, column=5, value=round(anom['deviation_from_mean_pct'], 1))
            ws.cell(row=row_idx, column=6, value=SEVERITY_LABELS.get(anom['severity'], anom['severity']))
            row_idx += 1

        self._autofit_columns(ws, len(headers), width=16)

    # ------------------------------------------------------------------
    # PDF Export
    # ------------------------------------------------------------------

    def export_pdf(self, filepath: str, title: str = "Reporte de Análisis Anual EPMaps") -> str:
        """Export comprehensive report to PDF document.

        Args:
            filepath: Path to save the .pdf file.
            title: Report title (default: "Reporte de Análisis Anual EPMaps").

        Returns:
            Path to the saved file.
        """
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontSize=22,
            textColor=colors.HexColor(f"#{HEADER_FILL}"),
            alignment=TA_CENTER,
            spaceAfter=6,
        )
        heading_style = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            textColor=colors.HexColor(f"#{HEADER_FILL}"),
            spaceBefore=18,
            spaceAfter=8,
        )
        body_style = styles["BodyText"]

        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(
            Paragraph(
                f"Período: Meses {self.analyzer.months[0]}-{self.analyzer.months[-1]}",
                ParagraphStyle("Subtitle", parent=body_style, alignment=TA_CENTER, textColor=colors.grey),
            )
        )
        elements.append(Spacer(1, 0.3 * inch))

        elements.extend(self._pdf_summary_section(heading_style))
        elements.extend(self._pdf_monthly_section(heading_style))
        elements.extend(self._pdf_categories_section(heading_style))

        if self.forecaster is not None:
            elements.append(PageBreak())
            elements.extend(self._pdf_forecast_section(heading_style))

        if self.detector is not None:
            elements.extend(self._pdf_anomalies_section(heading_style))

        doc.build(elements)
        return filepath

    def _pdf_table_style(self, n_rows: int) -> TableStyle:
        """Standard table style for PDF tables."""
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for row in range(1, n_rows):
            if row % 2 == 0:
                style.append(("BACKGROUND", (0, row), (-1, row), colors.HexColor(f"#{LIGHT_FILL}")))
        return TableStyle(style)

    def _pdf_summary_section(self, heading_style) -> list:
        """Build Resumen Anual section for PDF."""
        metrics = self.analyzer.yearly_metrics()
        elements = [Paragraph("Resumen Anual", heading_style)]

        data = [
            ["Métrica", "Valor"],
            ["Ingresos Totales", f"${metrics['total_revenue']:,.0f}"],
            ["Ingreso Promedio Mensual", f"${metrics['avg_monthly_revenue']:,.0f}"],
            ["Coeficiente de Variación", f"{metrics['cv_monthly_revenue']:.1f}%"],
            ["Mes Pico", f"{metrics['peak_month']} (${metrics['max_monthly_revenue']:,.0f})"],
            ["Mes Bajo", f"{metrics['low_month']} (${metrics['min_monthly_revenue']:,.0f})"],
            ["Registros Totales", f"{metrics['total_records']:,}"],
            ["Facturas Totales", f"{metrics['total_invoices']:,}"],
            ["Líneas Promedio por Factura", f"{metrics['avg_lines_per_invoice']:.2f}"],
        ]

        table = Table(data, colWidths=[3 * inch, 3 * inch])
        table.setStyle(self._pdf_table_style(len(data)))
        elements.append(table)
        return elements

    def _pdf_monthly_section(self, heading_style) -> list:
        """Build Progresión Mensual section for PDF."""
        metrics = self.analyzer.yearly_metrics()['monthly_metrics']
        changes = self.analyzer.monthly_changes()

        elements = [Paragraph("Progresión Mensual", heading_style)]

        data = [["Mes", "Ingresos", "Registros", "Precio Prom.", "Cambio %"]]
        for month in self.analyzer.months:
            m = metrics[month]
            change = changes[month].get('revenue_change')
            change_str = f"{change:+.1f}%" if change is not None else "—"
            data.append([
                month,
                f"${m['revenue']:,.0f}",
                f"{m['records']:,}",
                f"${m['avg_price']:.2f}",
                change_str,
            ])

        table = Table(data, colWidths=[0.8 * inch, 1.5 * inch, 1.3 * inch, 1.1 * inch, 1 * inch])
        table.setStyle(self._pdf_table_style(len(data)))
        elements.append(table)
        return elements

    def _pdf_categories_section(self, heading_style) -> list:
        """Build Top 5 Categorías section for PDF."""
        top_cats = self.analyzer.top_categories_yearly(top_n=5)

        elements = [Paragraph("Top 5 Categorías", heading_style)]

        data = [["Categoría", "Ingresos Totales", "% del Total", "Variabilidad"]]
        for rubro, cat_data in top_cats.items():
            data.append([
                rubro,
                f"${cat_data['total_revenue']:,.0f}",
                f"{cat_data['pct_of_total']:.1f}%",
                f"{cat_data['cv_monthly']:.1f}%",
            ])

        table = Table(data, colWidths=[1.5 * inch, 2 * inch, 1.5 * inch, 1.5 * inch])
        table.setStyle(self._pdf_table_style(len(data)))
        elements.append(table)
        return elements

    def _pdf_forecast_section(self, heading_style) -> list:
        """Build Pronóstico Q1 y Tendencias section for PDF."""
        forecast = self.forecaster.get_forecast_summary(periods=3)
        trends = self.forecaster.get_trends()

        elements = [Paragraph("Pronóstico Q1 y Tendencias", heading_style)]

        trend_label = TREND_LABELS.get(trends['trend_type'], trends['trend_type'].upper())
        elements.append(Paragraph(
            f"Tendencia: <b>{trend_label}</b> "
            f"({trends['annual_pct_change']:+.1f}% de cambio anual) | "
            f"Volatilidad: {trends['volatility_cv']:.1f}% | "
            f"Estacionalidad: {trends['seasonality_strength']:.1f}%",
            getSampleStyleSheet()["BodyText"],
        ))
        elements.append(Spacer(1, 0.15 * inch))

        data = [["Mes", "Pronóstico", "Límite Inferior", "Límite Superior"]]
        for month, fdata in forecast['revenue_forecast'].items():
            data.append([
                month,
                f"${fdata['forecast']:,.0f}",
                f"${fdata['lower_bound']:,.0f}",
                f"${fdata['upper_bound']:,.0f}",
            ])

        table = Table(data, colWidths=[1 * inch, 1.8 * inch, 1.8 * inch, 1.8 * inch])
        table.setStyle(self._pdf_table_style(len(data)))
        elements.append(table)

        elements.append(Spacer(1, 0.1 * inch))
        elements.append(Paragraph(
            f"Total Proyectado: <b>${forecast['projected_total_revenue']:,.0f}</b>",
            getSampleStyleSheet()["BodyText"],
        ))
        return elements

    def _pdf_anomalies_section(self, heading_style) -> list:
        """Build Detección de Anomalías section for PDF."""
        report = self.detector.get_overall_anomaly_report()
        alerts = self.detector.get_alerts(severity="all")

        elements = [Paragraph("Detección de Anomalías", heading_style)]

        risk_label = SEVERITY_LABELS.get(report['risk_level'], report['risk_level'].upper())
        elements.append(Paragraph(
            f"Nivel de Riesgo: <b>{risk_label}</b> "
            f"({report['risk_score']:.1f}/100) | "
            f"Anomalías Totales: {report['total_anomalies_detected']}",
            getSampleStyleSheet()["BodyText"],
        ))
        elements.append(Spacer(1, 0.15 * inch))

        if alerts:
            data = [["Tipo", "Mes", "Severidad", "Mensaje"]]
            for alert in alerts[:15]:
                data.append([
                    ALERT_TYPE_LABELS.get(alert['type'], alert['type'].capitalize()),
                    alert.get('month', '—'),
                    SEVERITY_LABELS.get(alert['severity'], alert['severity'].capitalize()),
                    alert['message'][:70] + ("..." if len(alert['message']) > 70 else ""),
                ])

            table = Table(data, colWidths=[0.9 * inch, 0.8 * inch, 0.9 * inch, 3.9 * inch])
            style = self._pdf_table_style(len(data))
            style.add("FONTSIZE", (0, 0), (-1, -1), 7.5)
            style.add("ALIGN", (3, 0), (3, -1), "LEFT")
            table.setStyle(style)
            elements.append(table)
        else:
            elements.append(Paragraph("No se detectaron anomalías significativas.", getSampleStyleSheet()["BodyText"]))

        return elements
